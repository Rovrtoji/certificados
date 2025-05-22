import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime

st.set_page_config(page_title="Procesador Certificados INTTELMEX", layout="centered")
st.title("📋 Procesador de Certificados INTTELMEX")

instaladores_file = st.file_uploader("📁 Subir archivo BASE INSTALADORES (.xlsx)", type="xlsx")
certificados_file = st.file_uploader("📁 Subir archivo BASE DE CERTIFICADOS INTTELMEX (.xlsx)", type="xlsx")

if st.button("🔄 Procesar archivos") and instaladores_file and certificados_file:
    try:
        wb1 = load_workbook(instaladores_file)
        ws1 = wb1.active

        df_certificados = pd.read_excel(certificados_file)
        df_certificados["Fin"] = pd.to_datetime(df_certificados["Fin"], errors='coerce')

        certificados_dict = {}
        for _, row in df_certificados.iterrows():
            nombre = str(row.get("NOMBRE (APELLIDOS + NOMBRES)", "")).strip().upper()
            if nombre not in certificados_dict:
                certificados_dict[nombre] = []
            certificados_dict[nombre].append(row)

        cert_cols = [9, 10, 11, 12, 13]
        nombre_coincidente_col = ws1.max_column + 1
        ws1.cell(row=1, column=nombre_coincidente_col).value = "Nombres Coincidentes"

        for row in range(2, ws1.max_row + 1):
            nombre1 = str(ws1.cell(row=row, column=2).value).strip().upper()
            nombre2 = str(ws1.cell(row=row, column=3).value).strip().upper()

            encontrados = []
            for nombre_clave in (nombre1, nombre2):
                if nombre_clave in certificados_dict:
                    encontrados = certificados_dict[nombre_clave]
                    break

            if encontrados:
                df_encontrados = pd.DataFrame(encontrados)
                df_encontrados = df_encontrados.dropna(subset=["Fin"])
                df_encontrados = df_encontrados.sort_values("Fin", ascending=False).head(5)

                certificados = df_encontrados["NUMERO DE CERTIFICADO INTTELMEX"].fillna("").tolist()
                nombres_coincidentes = df_encontrados["NOMBRE (APELLIDOS + NOMBRES)"].fillna("").tolist()

                for i, cert in enumerate(certificados):
                    ws1.cell(row=row, column=cert_cols[i], value=cert)

                ws1.cell(row=row, column=nombre_coincidente_col, value=" | ".join(nombres_coincidentes))

        output = BytesIO()
        wb1.save(output)
        st.success("✅ Archivo procesado exitosamente")
        st.download_button("⬇️ Descargar archivo actualizado", data=output.getvalue(),
                           file_name="BASE_INSTALADORES_ACTUALIZADO.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        st.error(f"❌ Error al procesar archivos: {e}")
